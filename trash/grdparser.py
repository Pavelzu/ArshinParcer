from selenium import webdriver
import time
import json
import re
from functools import singledispatch
import math
from datetime import datetime
import datetime
from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Alignment
from selenium.webdriver.chrome.options import Options
import sys
import os
import logging


def getpagesourcecode(link):
    options = Options()
    options.add_argument('--headless=new')
    # logger.info('Г')
    # driver = webdriver.Chrome(executable_path='/usr/bin/chromedriver',options=options)
#    driver = webdriver.Chrome(options=options)
    driver = webdriver.Chrome()

    driver.get(link)
    s = driver.page_source
    # time.sleep(10)
    driver.close()
    return s


@singledispatch
def getresultsquantity(resp):  # определение количества результатов выдачи
    subjson = resp['response']['numFound']
    return subjson


@getresultsquantity.register(str)
def _(page):  # определение количества результатов выдачи
    part = re.findall(r'"numFound":\d*', page)
    part = part[0].replace('"numFound":', '')
    return part


def get1stdownloadedpagequality(page, logger):
    isPageOK = 1
    if "response" in page and "responseHeader" in page and "numFound" in page:
        # print("Главная страница ОК")
        logger.info("Главная страница ОК")
    else:
        # print("Загрузка главной страницы пошла по бороде")
        logger.info("Загрузка главной страницы пошла по бороде")
        isPageOK = 0
    return isPageOK


def get2nddownloadedpagequality(page):
    isPageOK = 1
    if "result" in page and "miInfo" in page:
        # print("Инфо страница ОК")
        print("Инфо страница ОК")
    else:
        # print("Загрузка страницы c информацией пошла по бороде")

        isPageOK = 0
    return isPageOK


def removehtmltags(page):
    page = page.replace("</pre></body></html>", "")
    respbegin = page.find('{"result":')
    page = page[respbegin:len(page)]
    page = page.strip()
    return page


def clearindexes(blackindexes):
    whiteindexes = []
    for ind in blackindexes:
        tmpind = ind.replace('"vri_id":"', '')
        tmpind = tmpind.replace('"', '')
        whiteindexes.append(tmpind)
    return whiteindexes


def getinfofrominfopage(ind):
    page = ""
    url = "https://fgis.gost.ru/fundmetrology/cm/iaux/vri/" + str(ind)

    raknagore = 0
    while raknagore == 0:
        page = getpagesourcecode(url)
        raknagore = get2nddownloadedpagequality(page)
        # time.sleep(1)
    return page


def textUrlCombine(jsn, txt, url):
    jsn[txt] = "txt" + jsn[txt] + "url" + jsn[url]
    del jsn[url]
    return jsn


def jsontomonolist(jsn):
    result = []
    for key in jsn:
        result.append(key)
        result.append(jsn[key])
    return result


def jsoncleaner(jsn, whitedict):  # очистка json от лишних пар (не входят в итоговую выдачу)
    toDeleteList = []
    # print(jsn)
    for key in jsn:
        if key not in whitedict:
            toDeleteList.append(key)
    for itm in toDeleteList:
        del jsn[itm]
    # print(toDeleteList)
    result = dict()
    for key in jsn:
        result.update({whitedict[key]: jsn[key]})
    return jsontomonolist(result)  # возвращает лист из последовательностй поле, значение, поле, значение


def truefalsetranslater(jsn, booldict):
    for key in jsn:
        if str(jsn[key]).lower() == "false":
            jsn[key] = booldict["false"]
        if str(jsn[key]).lower() == "true":
            jsn[key] = booldict["true"]
    return jsn


def subjsonPoverkaInfoTranslator(jsn, vriTypekeys):
    for key in jsn:
        if str(jsn[key]).lower() == "1":
            jsn[key] = vriTypekeys["1"]
        if str(jsn[key]).lower() == "2":
            jsn[key] = vriTypekeys["2"]
    return jsn


def grandresponceparcer(urlind, jsonresp, targetkeys, vriTypekeys, boolkeys):
    if 'singleMI' in jsonresp['result']['miInfo']:
        subjsonSIInfo = jsonresp['result']['miInfo']['singleMI']  # Сведения о результатах поверки СИ
    if 'etaMI' in jsonresp['result']['miInfo']:
        subjsonSIInfo = jsonresp['result']['miInfo']['etaMI']  # Сведения о результатах поверки СИ
    subjsonSIInfo = textUrlCombine(subjsonSIInfo, "mitypeNumber", "mitypeURL")  # объединение номера СИ и url
    subjsonSIInfo.update({'mainurl': 'txtСсылкаurl' + 'https://fgis.gost.ru/fundmetrology/cm/results/' + urlind})
    subjsonPoverkaInfo = jsonresp['result']['vriInfo']
    try:
        subjsonApplicableInfo = jsonresp['result']['vriInfo']['applicable']
        del subjsonPoverkaInfo['applicable']
    except Exception:
        subjsonApplicableInfo = jsonresp['result']['vriInfo']['inapplicable']
        del subjsonPoverkaInfo['inapplicable']

    subjsonPoverkaInfo = {**subjsonPoverkaInfo, **subjsonApplicableInfo}  # Сведения о поверке
    subjsonPoverkaInfo = subjsonPoverkaInfoTranslator(subjsonPoverkaInfo,
                                                      vriTypekeys)  # перевод в нормальные слова Периодическая/первичная
    subjsonPoverkaInfo.update(checkPrigodnoSi(subjsonPoverkaInfo))  # определение пригодно ли СИ
    subjsonPoverkaInfo = truefalsetranslater(subjsonPoverkaInfo, boolkeys)
    subjsonDopSved = jsonresp['result']['info']  # Доп. сведения
    subjsonDopSved = truefalsetranslater(subjsonDopSved, boolkeys)
    '''
    Средства поверки Но пока не требуются

    jsonSredPoverBlocks = jsonresp['result']['means'] #В блоке "Средства поверки" всегда разные блоки в блоке means. Выявляем сперва их
    jsonsredProverList = []#Средства поверки ЛИСТ СЛОВАРЕЙ
    for key in jsonSredPoverBlocks:
        #targetkeys.update({key+"mitypeTitle": "Средства поверки"+key}) #на случай добавления инфы о средствах поверки в итоговый эксель
        if key != "uve":
            for jsn in jsonresp['result']['means'][key]:
                jsn = textUrlCombine(jsn, "mitypeTitle", "mitypeURL")
                jsn[key+'mitypeTitle'] = jsn.pop('mitypeTitle')
                jsonsredProverList.append(jsn)
            #print(jsn)
    #print(jsonsredProverList)
    '''
    infolist = []

    infolist += jsoncleaner(subjsonSIInfo, targetkeys)
    infolist += jsoncleaner(subjsonPoverkaInfo, targetkeys)
    infolist += jsoncleaner(subjsonDopSved, targetkeys)
    return infolist


def checkPrigodnoSi(jsn):
    result = {"SiPrigodno": "no info"}
    curdate = datetime.date.today()  # .isoformat()
    try:
        validDate = datetime.datetime.strptime(jsn["validDate"], "%d.%m.%Y").date()
        if curdate <= validDate:
            result["SiPrigodno"] = "Да"
        else:
            result["SiPrigodno"] = "Нет"
    except KeyError:
        result["SiPrigodno"] = "Нет"
    return result


def getsheetheader(data):  # формирование списка заголоска таблицы из неповторяющихся элементов
    headlist = []
    for card in data:
        i = 0
        for element in card:
            # print(element)
            if i % 2 != 0:
                i += 1
                continue
            else:
                if element not in headlist:
                    headlist.append(element)
            i += 1
    return headlist


def elementsplitter(el):  # разбивка значений типа txtблаблаurlблабла
    el = el.replace("txt", "")
    result = el.split("url")
    return result


def dateplustime():
    dt = datetime.datetime.now()
    date1st = dt.strftime('%d-%m-%Y')
    date2nd = dt.strftime('-%H-%M')
    return date1st + date2nd


def savetoexcel(biglist, filters, logger):
    header = getsheetheader(biglist)

    filename = makeFileName(filters)
    # создаем новую таблицу
    workbook = Workbook()
    sheet = workbook.active

    c = 1
    for el in header:  # добавляем заголовок в таблицу
        sheet.cell(row=1, column=c).value = el
        sheet.cell(row=1, column=c).alignment = Alignment(horizontal="center", vertical="center")
        c += 1

    r = 2
    for card in biglist:  # запись всех значений построчно в нужные столбцы
        i = 0
        for element in card:
            if i % 2 == 0:
                c = header.index(element) + 1
            else:
                try:
                    if "txt" in element and "url" in element:  # случай спайки текста и ссылки # сделать try
                        sheet.cell(row=r, column=c).hyperlink = elementsplitter(element)[1]
                        sheet.cell(row=r, column=c).value = elementsplitter(element)[0]
                        sheet.cell(row=r, column=c).style = "Hyperlink"
                        # sheet.cell(row=r, column=c).value = element

                    else:
                        if "https" in element:  # если только ссылка в значении
                            sheet.cell(row=r, column=c).hyperlink = element
                            sheet.cell(row=r, column=c).value = element
                            sheet.cell(row=r, column=c).style = "Hyperlink"
                        else:
                            sheet.cell(row=r, column=c).value = element
                except Exception:
                    sheet.cell(row=r,
                               column=c).value = element  # только ради года выпуска СИ, который приходил числом и говнился в if выше
            i += 1
        r += 1

    for i in range(1, 30):  # свистоперделки с шириной столбца
        letter = get_column_letter(i)
        sheet.column_dimensions[letter].width = 25
    workbook.save(filename="./results/" + filename)
    logger.info('Записали в файл <a href = "/results/' + filename + '">' + filename + '</a>')


def makeMainLink(arguments, rows):
    filters = arguments
    filters = filters.replace('|', '"')
    # https://fgis.gost.ru/fundmetrology/cm/xcdb/vri/select?fq=verification_year:2024&fq=org_title:*1*&fq=mi.mitnumber:*2*&fq=mi.mititle:*3*&fq=mi.mitype:*4*&fq=mi.modification:*5*&fq=verification_date:[2024-09-05T00:00:00Z%20TO%202024-09-20T23:59:59Z]&q=*&fl=vri_id,org_title,mi.mitnumber,mi.mititle,mi.mitype,mi.modification,mi.number,verification_date,valid_date,applicability,result_docnum,sticker_num&sort=verification_date+desc,org_title+asc&rows=20&start=0
    reqHead = 'https://fgis.gost.ru/fundmetrology/cm/xcdb/vri/select?'
    reqTail = 'q=*&fl=vri_id,org_title,mi.mitnumber,mi.mititle,mi.mitype,mi.modification,mi.number,verification_date,valid_date,applicability,result_docnum,sticker_num&sort=verification_date+desc,org_title+asc&rows=' + str(
        rows)  # вырезан &start=0
    filtersDict = json.loads(filters)

    reqBody = ''

    if filtersDict["verification_year"] != "":
        reqBody = reqBody + "fq=verification_year:" + filtersDict["verification_year"] + "&"

    for filter in filtersDict:
        if filtersDict[filter] == "" or filter == "verification_year":
            continue
        if filter != "verification_date":
            curentFilterValue = filtersDict[filter]
            curentFilterValue = curentFilterValue.replace('^', '\\"')
            curentFilterValue = curentFilterValue.replace('-', '\\-')
            singleWords = curentFilterValue.split()
            for word in singleWords:
                reqBody = reqBody + "fq=" + filter + ":*" + word + "*&"
        else:
            reqBody = reqBody + "fq=" + filter + ":" + filtersDict[
                filter] + "&"  # диапазон дат вставляется без * по бокам

    fullReq = reqHead + reqBody + reqTail
    return fullReq


def makeFileName(arguments):
    badSymbols = {ord('+'): None, ord('='): None, ord('.'): None, ord('['): None, ord(':'): None, ord(']'): None,
                  ord('*'): None, ord('?'): None, ord(';'): None, ord('.'): None, ord('|'): None, ord('\\'): None,
                  ord('/'): None}
    filters = arguments
    filters = filters.replace('|', '"')
    filtersDict = json.loads(filters)
    fn = ''

    if filtersDict["mi.mititle"]:
        fn = fn + filtersDict["mi.mititle"] + '_'
    if filtersDict["mi.mitnumber"]:
        fn = fn + filtersDict["mi.mitnumber"] + '_'
    if filtersDict["mi.mitype"]:
        fn = fn + filtersDict["mi.mitype"] + '_'
    if filtersDict["mi.modification"]:
        fn = fn + filtersDict["mi.modification"] + '_'
    if filtersDict["org_title"]:
        fn = fn + filtersDict["org_title"] + '_'

    fn = fn.translate(badSymbols)
    # fn = fn.replace('^','\"')
    fn = fn + "_" + dateplustime() + ".xlsx"

    return fn

def main():
    dati = dateplustime()
    print('Ищем ' + sys.argv[1])
    pid = os.getpid()
    with open("pidlogargs.txt", "a") as file:
        file.write(str(pid) + ' ' + dati + '.log ' + sys.argv[1]+'\n')

    # logger.info('Пишем в pidlogargs ' + str(pid) + ' ' + dati + '.log ' + sys.argv[1])
    grandinfolist = []
    targetkeys = {'mainurl': 'Исходник', 'mitypeNumber': 'Рег номер типа СИ', 'mitypeType': 'Тип СИ',
                  'mitypeTitle': 'Наименование типа СИ',
                  'manufactureNum': 'Заводской номер СИ', 'regNumber': "Регистрационный номер СИ в реестре ФИФ ОЕИ",
                  'manufactureYear': 'Год выпуска СИ',
                  'modification': 'Модификация СИ', 'organization': 'Наименование организации-поверителя',
                  'signCipher': 'Условный шифр знака поверки', 'miOwner': 'Владелец СИ', 'vrfDate': 'Дата поверки СИ',
                  'rankCode': 'Код разряда эталона в ГПС',
                  'validDate': 'Поверка действительна до', 'docTitle': 'Наименование документа, на основании'
                                                                       'которого выполнена поверка',
                  'rankTitle': 'Наименование разряда эталона в ГПС/ЛПС', 'vriType': 'Тип поверки',
                  'certNum': 'Номер свидетельства', 'noticeNum': 'Номер извещения', 'signPass':
                      'Знак поверки в паспорте', 'signMi': 'Знак поверки на СИ',
                  'briefIndicator': 'Поверка в сокращенном объеме', "SiPrigodno": "СИ пригодно",
                  "additional_info": "Доп сведения", 'structure':'Состав СИ, представленного на поверку'}
    boolkeys = {"false": "Нет", "true": "Да"}
    vriTypekeys = {'1': 'Первичная', '2': 'Периодическая'}
    ind = '1-373123835'  #https://fgis.gost.ru/fundmetrology/cm/iaux/vri/1-373123835
    curinfolist = getinfofrominfopage(ind)
    jsonresptext = removehtmltags(curinfolist)
    jsonresp = json.loads(jsonresptext)
    curinfolist = grandresponceparcer(ind, jsonresp, targetkeys, vriTypekeys, boolkeys)  # urlind, jsonresp, targetkeys, boolkeys

    print (curinfolist)

if __name__ == "__main__":
    main()